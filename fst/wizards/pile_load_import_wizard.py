from odoo import models, fields, api
from odoo.exceptions import ValidationError
from openpyxl import load_workbook
import base64
import io
from datetime import datetime, time
from pytz import timezone
import pytz

india_tz = timezone('Asia/Kolkata')

MODEL_CONFIG = {
    "pile.load.test.parent": {
        "loading": "pile.load.reading.loading",
        "unloading": "pile.load.reading.unloading",
        "dials": 4,
    },
    "routine.pile.load.test": {
        "loading": "routine.pile.reading.loading",
        "unloading": "routine.pile.reading.unloading",
        "dials": 4,
    },
    "pullout.pile.load.test.parent": {
        "loading": "pullout.pile.load.reading.loading",
        "unloading": "pullout.pile.load.reading.unloading",
        "dials": 2,
    },
    "routine.pullout.pile.load.test.parent": {
        "loading": "routine.pullout.pile.load.reading.loading",
        "unloading": "routine.pullout.pile.load.reading.unloading",
        "dials": 2,
    },
    "lateral.pile.load.test.parent": {
        "loading": "lateral.pile.load.reading.loading",
        "unloading": "lateral.pile.load.reading.unloading",
        "dials": 2,
    },
    "routine.lateral.pile.load.test.parent": {
        "loading": "routine.lateral.pile.load.reading.loading",
        "unloading": "routine.lateral.pile.load.reading.unloading",
        "dials": 2,
    },
}

class PileLoadImportWizard(models.TransientModel):
    _name = 'pile.load.import.wizard'
    _description = 'Pile Load Excel Import Wizard'

    file = fields.Binary(required=True)
    filename = fields.Char()

    parent_model = fields.Char()
    parent_id = fields.Integer()

    def action_import_excel(self):

        self.ensure_one()
        # import wdb;wdb.set_trace()

        if not self.file:
            raise ValidationError("Please upload an Excel file.")

        file_data = base64.b64decode(self.file)

        workbook = load_workbook(
            filename=io.BytesIO(file_data),
            data_only=True
        )

        required_sheets = ['Loading', 'Unloading']

        for sheet_name in required_sheets:
            if sheet_name not in workbook.sheetnames:
                raise ValidationError(
                    f"Sheet '{sheet_name}' not found in Excel."
                )

        config = MODEL_CONFIG.get(self.parent_model)
        parent = self.env[self.parent_model].browse(self.parent_id)

        if not parent.exists():
            raise ValidationError("Parent record not found.")

        if not config:
            raise ValidationError("Unsupported parent model.")

        self._process_sheet(
            workbook['Loading'],
            config["loading"],
            config["dials"]
        )

        self._process_sheet(
            workbook['Unloading'],
            config["unloading"],
            config["dials"],
        )

        for method in (
            "action_recompute_all",
            "action_generate_graph",
            "action_generate_analysis",
            "_compute_qr_code",
        ):
            if hasattr(parent, method):
                getattr(parent, method)()
            

    def _process_sheet(self, sheet, model_name,dials):

        Model = self.env[model_name]

        headers = []
        if dials == 4:
            headers = [
                'Date',
                'Time',
                'Load',
                'Dial A',
                'Dial B',
                'Dial C',
                'Dial D',
            ]
        else:
            headers = [
                'Date',
                'Time',
                'Load',
                'Dial A',
                'Dial B',
            ]

        for index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            row_data = dict(zip(headers, row))
            # Skip fully empty rows
            # if not any(row):
            #     continue

            required = [
                "Load",
                "Dial A",
                "Dial B",
            ]

            if dials == 4:
                required.extend(["Dial C", "Dial D"])

            if all(row_data[h] in (None, "") for h in required):
                continue

            for field_name, value in row_data.items():

                if value in (None, ''):
                    raise ValidationError(
                        f"{field_name} is empty in "
                        f"{sheet.title} sheet row {index}"
                    )

            excel_date = row_data['Date']
            excel_time = row_data['Time']

            
            if not isinstance(excel_date, datetime):
                raise ValidationError(
                    f"Invalid Date in "
                    f"{sheet.title} sheet row {index}"
                )

            if isinstance(excel_time, datetime):
                time_part = excel_time.time()

            elif isinstance(excel_time, time):
                time_part = excel_time

            else:
                raise ValidationError(
                    f"Invalid Time in "
                    f"{sheet.title} sheet row {index}"
                )

            excel_dt = datetime.combine(
                excel_date.date(),
                time_part
            )

            # Excel times are local India times
            localized_dt = india_tz.localize(excel_dt)

            # Convert to UTC for Odoo storage
            normalized_dt = localized_dt.astimezone(
                pytz.UTC
            ).replace(
                tzinfo=None,
                second=0,
                microsecond=0
            )

            # existing = Model.search([
            #     ("parent_id", "=", self.parent_id)
            # ])

            matched_record = Model.search([
                ("parent_id", "=", self.parent_id),
                ("load_tonne", "=", row_data["Load"]),
                ("reading_datetime", "=", normalized_dt),
            ], limit=1)

            vals = {
                'parent_id': self.parent_id,
                'reading_datetime': normalized_dt,
                'load_tonne': row_data['Load'],
                'dial_a': row_data['Dial A'],
                'dial_b': row_data['Dial B'],
            }
            if dials == 4:
                vals.update({
                    "dial_c": row_data["Dial C"],
                    "dial_d": row_data["Dial D"],
                })

            if matched_record:
                matched_record.write(vals)
            else:
                Model.create(vals)

