# Preview conversion helpers (LibreOffice headless + HTML fallbacks).

import csv
import html
import os
import shutil
import subprocess
import tempfile
from urllib.parse import quote

from openpyxl import load_workbook

import config


class ConversionError(Exception):
    pass


def get_or_create_pdf(src_path):
    """Return a cached PDF rendering of the source document (docx/doc/xlsx/xls)."""
    sha = config.sha256_of(src_path)
    cache_path = os.path.join(config.PREVIEW_CACHE_DIR, sha + '.pdf')
    if os.path.exists(cache_path):
        return cache_path
    with tempfile.TemporaryDirectory(prefix='dms_conv_') as tmp:
        pdf_path = _libreoffice_pdf(src_path, tmp)
        shutil.move(pdf_path, cache_path)
    return cache_path


def _libreoffice_pdf(src_path, out_dir):
    base = os.path.splitext(os.path.basename(src_path))[0]
    out_file = os.path.join(out_dir, base + '.pdf')
    cmd = [
        config.LIBREOFFICE_BIN, '--headless', '--norestore',
        '--convert-to', 'pdf', '--outdir', out_dir, src_path,
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=240)
    except FileNotFoundError as exc:
        raise ConversionError(
            'LibreOffice not found. Install it or set DMS_LIBREOFFICE_BIN.') from exc
    except subprocess.TimeoutExpired as exc:
        raise ConversionError('Conversion timed out.') from exc
    if result.returncode != 0 or not os.path.exists(out_file):
        raise ConversionError('LibreOffice conversion failed: %s' % result.stderr.decode()[:300])
    return out_file


def _read_text(path):
    for encoding in ('utf-8', 'utf-8-sig', 'latin-1'):
        try:
            with open(path, 'r', encoding=encoding, newline='') as f:
                return f.read()
        except (UnicodeDecodeError, OSError):
            continue
    return ''


def csv_to_html(path, max_rows=500):
    content = _read_text(path)
    reader = csv.reader(content.splitlines())
    rows = []
    for idx, row in enumerate(reader):
        if idx >= max_rows:
            break
        rows.append(row)
    if not rows:
        return '<p class="dms-empty">Empty CSV file.</p>'
    head = rows[0]
    body = rows[1:]
    tr = lambda cells: '<tr>' + ''.join('<td>%s</td>' % html.escape(str(c)) for c in cells) + '</tr>'
    thead = '<thead><tr>' + ''.join('<th>%s</th>' % html.escape(str(c)) for c in head) + '</tr></thead>'
    tbody = '<tbody>' + ''.join(tr(r) for r in body) + '</tbody>'
    return (
        '<div class="dms-table-wrap">'
        '<table class="dms-preview-table">%s%s</table>'
        '</div>'
        '<p class="dms-preview-note">Showing up to %d rows.</p>' % (thead, tbody, max_rows)
    )


def xlsx_to_html(path, max_rows=500):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= max_rows:
            break
        rows.append([v if v is not None else '' for v in row])
    wb.close()
    if not rows:
        return '<p class="dms-empty">Empty spreadsheet.</p>'
    head = rows[0]
    body = rows[1:]
    tr = lambda cells: '<tr>' + ''.join('<td>%s</td>' % html.escape(str(c)) for c in cells) + '</tr>'
    thead = '<thead><tr>' + ''.join('<th>%s</th>' % html.escape(str(c)) for c in head) + '</tr></thead>'
    tbody = '<tbody>' + ''.join(tr(r) for r in body) + '</tbody>'
    return (
        '<div class="dms-table-wrap">'
        '<table class="dms-preview-table">%s%s</table>'
        '</div>'
        '<p class="dms-preview-note">Showing up to %d rows.</p>' % (thead, tbody, max_rows)
    )


def guess_kind(path):
    name = os.path.basename(path).lower()
    ext = os.path.splitext(name)[1]
    if ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg', '.bmp'):
        return 'image'
    if ext == '.pdf':
        return 'pdf'
    if ext in ('.docx', '.doc'):
        return 'word'
    if ext in ('.xlsx', '.xls'):
        return 'excel'
    if ext == '.csv':
        return 'csv'
    return 'other'
