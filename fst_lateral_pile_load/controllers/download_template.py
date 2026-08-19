import base64
import io
from datetime import datetime, timedelta

from odoo import http
from odoo.http import request


class LateralPileLoadTemplateController(http.Controller):

    @http.route(
        "/fst_lateral_pile_load/header_image/<int:lab_id>",
        type="http",
        auth="public",
        csrf=False,
    )
    def header_image(self, lab_id):
        lab = request.env["lerm.lab.master"].sudo().browse(lab_id)
        data = lab.header_image
        if not data:
            return request.not_found()
        raw = base64.b64decode(data) if isinstance(data, bytes) else data
        attachment = request.env["ir.attachment"].sudo().search([
            ("res_model", "=", "lerm.lab.master"),
            ("res_field", "=", "header_image"),
            ("res_id", "=", lab_id),
        ], limit=1)
        mimetype = attachment.mimetype or "image/png"
        return request.make_response(
            raw,
            headers=[("Content-Type", mimetype)],
        )

    @http.route(
        "/fst_lateral_pile_load/download_template",
        type="http",
        auth="user",
    )
    def download_template(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        except ImportError:
            return request.make_response(
                b"openpyxl library is not available.",
                headers=[("Content-Type", "text/plain")],
            )

        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="244061", end_color="244061", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        data_font = Font(size=10)
        data_align = Alignment(horizontal="center")
        dt_align = Alignment(horizontal="center")
        dt_fmt = "DD/MM/YYYY HH:MM"

        def fill_sheet(ws, title, sample_data):
            ws.title = title
            headers = ["Date/Time", "Load (Tonne)", "Dial A (mm)", "Dial B (mm)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18

            for r, (dt_val, load, a, b) in enumerate(sample_data, 2):
                cell_dt = ws.cell(row=r, column=1, value=dt_val)
                cell_dt.font = data_font
                cell_dt.alignment = dt_align
                cell_dt.number_format = dt_fmt

                ws.cell(row=r, column=2, value=load).font = data_font
                ws.cell(row=r, column=3, value=a).font = data_font
                ws.cell(row=r, column=4, value=b).font = data_font
                for c in range(1, 5):
                    ws.cell(row=r, column=c).border = thin_border
                    if c > 1:
                        ws.cell(row=r, column=c).alignment = data_align

        base_dt = datetime.now().replace(hour=8, minute=0, second=0, microsecond=0)

        loading_sample = [
            (base_dt, 0.0, 0.00, 0.00),
            (base_dt + timedelta(minutes=15), 22.5, 0.52, 0.48),
            (base_dt + timedelta(minutes=30), 22.5, 0.85, 0.80),
            (base_dt + timedelta(minutes=45), 22.5, 1.05, 1.00),
            (base_dt + timedelta(hours=1), 45.0, 1.40, 1.35),
            (base_dt + timedelta(hours=1, minutes=15), 45.0, 1.70, 1.65),
            (base_dt + timedelta(hours=1, minutes=30), 45.0, 1.90, 1.85),
            (base_dt + timedelta(hours=1, minutes=45), 67.5, 2.25, 2.20),
            (base_dt + timedelta(hours=2), 67.5, 2.55, 2.50),
            (base_dt + timedelta(hours=2, minutes=15), 67.5, 2.75, 2.70),
            (base_dt + timedelta(hours=2, minutes=30), 90.0, 3.10, 3.05),
            (base_dt + timedelta(hours=2, minutes=45), 90.0, 3.40, 3.35),
            (base_dt + timedelta(hours=3), 90.0, 3.60, 3.55),
            (base_dt + timedelta(hours=3, minutes=15), 112.5, 4.00, 3.95),
            (base_dt + timedelta(hours=3, minutes=30), 112.5, 4.30, 4.25),
            (base_dt + timedelta(hours=3, minutes=45), 112.5, 4.50, 4.45),
            (base_dt + timedelta(hours=4), 135.0, 4.90, 4.85),
            (base_dt + timedelta(hours=4, minutes=15), 135.0, 5.20, 5.15),
            (base_dt + timedelta(hours=4, minutes=30), 135.0, 5.40, 5.35),
            (base_dt + timedelta(hours=4, minutes=45), 157.5, 5.80, 5.75),
            (base_dt + timedelta(hours=5), 157.5, 6.10, 6.05),
            (base_dt + timedelta(hours=5, minutes=15), 157.5, 6.30, 6.25),
            (base_dt + timedelta(hours=5, minutes=30), 180.0, 6.70, 6.65),
            (base_dt + timedelta(hours=5, minutes=45), 180.0, 7.00, 6.95),
            (base_dt + timedelta(hours=6), 180.0, 7.20, 7.15),
        ]
        unloading_sample = [
            (base_dt + timedelta(hours=6, minutes=15), 135.0, 6.20, 6.15),
            (base_dt + timedelta(hours=6, minutes=30), 135.0, 5.80, 5.75),
            (base_dt + timedelta(hours=6, minutes=45), 135.0, 5.50, 5.45),
            (base_dt + timedelta(hours=7), 90.0, 4.60, 4.55),
            (base_dt + timedelta(hours=7, minutes=15), 90.0, 4.20, 4.15),
            (base_dt + timedelta(hours=7, minutes=30), 90.0, 3.90, 3.85),
            (base_dt + timedelta(hours=7, minutes=45), 45.0, 2.80, 2.75),
            (base_dt + timedelta(hours=8), 45.0, 2.40, 2.35),
            (base_dt + timedelta(hours=8, minutes=15), 45.0, 2.10, 2.05),
            (base_dt + timedelta(hours=8, minutes=30), 0.0, 0.80, 0.75),
            (base_dt + timedelta(hours=8, minutes=45), 0.0, 0.65, 0.60),
        ]

        ws1 = wb.active
        fill_sheet(ws1, "Loading", loading_sample)

        ws2 = wb.create_sheet()
        fill_sheet(ws2, "Unloading", unloading_sample)

        output = io.BytesIO()
        wb.save(output)
        wb.close()

        return request.make_response(
            output.getvalue(),
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", "attachment; filename=Lateral_Pile_Load_Template.xlsx"),
            ],
        )
