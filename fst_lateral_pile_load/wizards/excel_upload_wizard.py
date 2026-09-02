import base64
import io
import re
from datetime import datetime, timedelta

from odoo import api, fields, models, _
from odoo.exceptions import UserError


class LateralPileLoadExcelUpload(models.TransientModel):
    _name = "lateral.pile.load.excel.upload"
    _description = "Excel Upload for Lateral Pile Load Readings"

    file_data = fields.Binary(string="Excel File")
    file_name = fields.Char(string="Filename")

    parent_id = fields.Many2one("fst.lateral.pile.load.test", string="Parent Record", required=True)

    def action_upload(self):
        self.ensure_one()

        if not self.file_data:
            raise UserError(_("Please select an Excel file to upload."))

        file_data = base64.b64decode(self.file_data)
        stream = io.BytesIO(file_data)

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("openpyxl library is not available."))

        try:
            wb = load_workbook(stream, data_only=False)
        except Exception as e:
            raise UserError(_("Could not read the Excel file: %s") % str(e))

        sheet_names = wb.sheetnames
        if len(sheet_names) < 2:
            raise UserError(
                _("Excel must have 2 sheets: 'Loading' and 'Unloading'. Found: %s")
                % ", ".join(sheet_names)
            )

        self._parse_sheet(wb, "Loading", "fst.lateral.pile.load.reading.loading")
        self._parse_sheet(wb, "Unloading", "fst.lateral.pile.load.reading.unloading")

        wb.close()

        self.parent_id.action_recompute_all()

        return {
            "type": "ir.actions.act_window_close",
        }

    def _parse_sheet(self, wb, sheet_name, model_name):
        if sheet_name not in wb.sheetnames:
            raise UserError(_("Sheet '%s' not found in the Excel file.") % sheet_name)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        existing = self.env[model_name].search([("parent_id", "=", self.parent_id.id)])
        existing.unlink()

        row_num = 0
        for row in rows[1:]:
            row_num += 1
            if not any(v is not None and v != "" for v in row):
                continue

            try:
                dt = self._parse_datetime(row[0])
                load_val = self._parse_float(row[1]) if len(row) > 1 else 0.0
                dial_a = self._parse_float(row[2]) if len(row) > 2 else 0.0
                dial_b = self._parse_float(row[3]) if len(row) > 3 else 0.0
                dial_c = self._parse_float(row[4]) if len(row) > 4 else 0.0

                vals = {
                    "parent_id": self.parent_id.id,
                    "reading_datetime": dt,
                    "load_tonne": load_val,
                    "applied_pressure": round(load_val * 1000.0 / 154.0, 2),
                    "pressure_under_plate": round(load_val / 0.07, 2),
                    "dial_a": dial_a,
                    "dial_b": dial_b,
                    "dial_c": dial_c,
                }

                self.env[model_name].create(vals)
            except Exception as e:
                raise UserError(
                    _("Error processing %s row %s: %s") % (sheet_name, row_num, str(e))
                )

    def _parse_datetime(self, val):
        if val is None:
            return fields.Datetime.now()
        if isinstance(val, datetime):
            return val
        if isinstance(val, str):
            val = val.strip()
            if not val:
                return fields.Datetime.now()
            for fmt in ["%d/%m/%Y %H:%M", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%y %H:%M"]:
                try:
                    return datetime.strptime(val, fmt)
                except ValueError:
                    continue
        return fields.Datetime.now()

    def _parse_float(self, val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            val = val.strip()
            if not val:
                return 0.0
            val = re.sub(r"[^\d.\-]", "", val)
            return float(val) if val else 0.0
        return 0.0

    def action_download_template(self):
        return {
            "type": "ir.actions.act_url",
            "url": "/fst_lateral_pile_load/download_template",
            "target": "new",
        }
