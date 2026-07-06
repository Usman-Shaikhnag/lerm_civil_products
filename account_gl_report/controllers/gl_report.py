import json
from datetime import datetime

from dateutil.relativedelta import relativedelta

from odoo import fields, http
from odoo.http import request


class GLReportController(http.Controller):

    def _get_report(self):
        return request.env['account.gl.report'].sudo()

    def _resolve_dates(self, params):
        date_filter = params.get('date_filter', 'this_month')
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        if date_filter != 'custom' or not (date_from and date_to):
            today = fields.Date.today()
            if date_filter == 'today':
                date_from = date_to = today
            elif date_filter == 'yesterday':
                date_from = date_to = today - relativedelta(days=1)
            elif date_filter == 'this_week':
                date_from = today + relativedelta(weekday=0, weeks=0)
                date_to = today
            elif date_filter == 'last_week':
                date_from = today + relativedelta(weekday=0, weeks=-1)
                date_to = today + relativedelta(weekday=6, weeks=-1)
            elif date_filter == 'this_month':
                date_from = today.replace(day=1)
                date_to = today
            elif date_filter == 'last_month':
                first_of_this = today.replace(day=1)
                date_to = first_of_this - relativedelta(days=1)
                date_from = date_to.replace(day=1)
            elif date_filter == 'quarter':
                quarter = (today.month - 1) // 3
                date_from = today.replace(month=quarter * 3 + 1, day=1)
                date_to = today
            elif date_filter == 'fiscal_year':
                company = request.env.company
                fy = company.compute_fiscalyear_dates(today)
                date_from = fy['date_from']
                date_to = fy['date_to']
        params['date_from'] = date_from
        params['date_to'] = date_to
        return params

    @http.route('/gl_report/get_data', type='json', auth='user', methods=['POST'])
    def get_data(self, **kwargs):
        params = dict(request.params)
        params = self._resolve_dates(params)
        report = self._get_report()
        group_by = params.get('group_by', 'none')

        if group_by and group_by != 'none':
            return report.get_grouped_data(params)
        return report._build_query(params)

    @http.route('/gl_report/get_line_details', type='json', auth='user', methods=['POST'])
    def get_line_details(self, line_id, **kwargs):
        report = self._get_report()
        return report.get_line_details(line_id)

    @http.route('/gl_report/get_attachments', type='json', auth='user', methods=['POST'])
    def get_attachments(self, move_id, **kwargs):
        report = self._get_report()
        return report.get_attachments(move_id)

    @http.route('/gl_report/search_accounts', type='json', auth='user', methods=['POST'])
    def search_accounts(self, term='', **kwargs):
        domain = []
        if term:
            domain = ['|', ('code', 'ilike', term), ('name', 'ilike', term)]
        accounts = request.env['account.account'].sudo().search_read(
            domain, ['id', 'code', 'name'], limit=20)
        return [{'id': a['id'], 'display_name': f"[{a['code']}] {a['name']}"}
                for a in accounts]

    @http.route('/gl_report/search_partners', type='json', auth='user', methods=['POST'])
    def search_partners(self, term='', **kwargs):
        domain = [('parent_id', '=', False)]
        if term:
            domain.append(('name', 'ilike', term))
        partners = request.env['res.partner'].sudo().search_read(
            domain, ['id', 'name'], limit=20)
        return [{'id': p['id'], 'display_name': p['name']} for p in partners]

    @http.route('/gl_report/search_analytic', type='json', auth='user', methods=['POST'])
    def search_analytic(self, term='', **kwargs):
        domain = []
        if term:
            domain.append(('name', 'ilike', term))
        accounts = request.env['account.analytic.account'].sudo().search_read(
            domain, ['id', 'name'], limit=20)
        return [{'id': a['id'], 'display_name': a['name']} for a in accounts]

    @http.route('/gl_report/get_journals', type='json', auth='user', methods=['POST'])
    def get_journals(self, **kwargs):
        journals = request.env['account.journal'].sudo().search_read(
            [], ['id', 'code', 'name'], order='code')
        return [{'id': j['id'], 'display_name': f"[{j['code']}] {j['name']}"}
                for j in journals]

    @http.route('/gl_report/save_preset', type='json', auth='user', methods=['POST'])
    def save_preset(self, name='', params_json='', **kwargs):
        preset = request.env['account.gl.report.preset'].sudo()
        return {'id': preset.save_preset(name, params_json)}

    @http.route('/gl_report/get_presets', type='json', auth='user', methods=['POST'])
    def get_presets(self, **kwargs):
        preset = request.env['account.gl.report.preset'].sudo()
        return {'presets': preset.get_user_presets()}

    @http.route('/gl_report/delete_preset', type='json', auth='user', methods=['POST'])
    def delete_preset(self, preset_id, **kwargs):
        request.env['account.gl.report.preset'].sudo().browse(preset_id).unlink()
        return {'success': True}

    @http.route('/gl_report/export', type='json', auth='user', methods=['POST'])
    def export(self, params=None, export_format='csv', **kwargs):
        params = self._resolve_dates(params or {})
        report = self._get_report()
        data = report._build_query(params)
        records = data['records']

        if export_format == 'csv':
            return self._export_csv(records, params)
        elif export_format == 'xlsx':
            return self._export_xlsx(records, params)
        elif export_format == 'pdf':
            return self._export_pdf(records, params)
        return {'error': 'Unsupported format'}

    def _export_csv(self, records, params):
        import csv
        import io
        import base64

        output = io.StringIO()
        writer = csv.writer(output)
        headers = ['Date', 'Journal', 'Move', 'Partner', 'Label',
                    'Reference', 'Debit', 'Credit', 'Balance',
                    'Analytic Account', 'State']
        writer.writerow(headers)
        for r in records:
            writer.writerow([
                r.get('date', ''),
                r.get('journal_code', ''),
                r.get('move_name', ''),
                r.get('partner_name', ''),
                r.get('label', ''),
                r.get('ref', ''),
                r.get('debit', 0),
                r.get('credit', 0),
                r.get('running_balance', r.get('balance', 0)),
                r.get('analytic_name', ''),
                r.get('state', ''),
            ])
        content = output.getvalue()
        filename = f"general_ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        attachment = request.env['ir.attachment'].sudo().create({
            'name': filename,
            'datas': base64.b64encode(content.encode('utf-8')),
            'mimetype': 'text/csv',
        })
        return {
            'url': f'/web/content/{attachment.id}?download=true',
            'filename': filename,
        }

    def _export_xlsx(self, records, params):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return {'error': 'openpyxl library not available'}

        import base64
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = 'General Ledger'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        headers = ['Date', 'Journal', 'Move', 'Partner', 'Label',
                    'Reference', 'Debit', 'Credit', 'Balance',
                    'Analytic Account', 'State']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, r in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=str(r.get('date', '')))
            ws.cell(row=row_num, column=2, value=r.get('journal_code', ''))
            ws.cell(row=row_num, column=3, value=r.get('move_name', ''))
            ws.cell(row=row_num, column=4, value=r.get('partner_name', ''))
            ws.cell(row=row_num, column=5, value=r.get('label', ''))
            ws.cell(row=row_num, column=6, value=r.get('ref', ''))
            c_debit = ws.cell(row=row_num, column=7, value=r.get('debit', 0))
            c_debit.number_format = '#,##0.00'
            c_credit = ws.cell(row=row_num, column=8, value=r.get('credit', 0))
            c_credit.number_format = '#,##0.00'
            c_bal = ws.cell(row=row_num, column=9, value=r.get('running_balance', r.get('balance', 0)))
            c_bal.number_format = '#,##0.00'
            ws.cell(row=row_num, column=10, value=r.get('analytic_name', ''))
            ws.cell(row=row_num, column=11, value=r.get('state', ''))

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        filename = f"general_ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        attachment = request.env['ir.attachment'].sudo().create({
            'name': filename,
            'datas': base64.b64encode(content),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'url': f'/web/content/{attachment.id}?download=true',
            'filename': filename,
        }

    def _export_pdf(self, records, params):
        import base64
        from io import BytesIO
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph(f"General Ledger Report", styles['Title']))
        elements.append(Spacer(1, 12))
        company = request.env.company
        elements.append(Paragraph(f"Company: {company.name}", styles['Normal']))
        elements.append(Spacer(1, 6))

        if records:
            headers = ['Date', 'Journal', 'Move', 'Partner', 'Label', 'Ref', 'Debit', 'Credit', 'Balance']
            table_data = [headers]
            for r in records:
                table_data.append([
                    str(r.get('date', '') or ''),
                    r.get('journal_code', '') or '',
                    r.get('move_name', '') or '',
                    r.get('partner_name', '') or '',
                    (r.get('label', '') or '')[:40],
                    (r.get('ref', '') or '')[:20],
                    f"{r.get('debit', 0):,.2f}",
                    f"{r.get('credit', 0):,.2f}",
                    f"{r.get('running_balance', r.get('balance', 0)):,.2f}",
                ])

            table = Table(table_data, repeatRows=1)
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (6, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ])
            table.setStyle(style)
            elements.append(table)

        doc.build(elements)
        content = buf.getvalue()
        filename = f"general_ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        attachment = request.env['ir.attachment'].sudo().create({
            'name': filename,
            'datas': base64.b64encode(content),
            'mimetype': 'application/pdf',
        })
        return {
            'url': f'/web/content/{attachment.id}?download=true',
            'filename': filename,
        }
